"""
Script to fix Excel column widths and enable text wrapping
This will make the RSA private key visible in Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def fix_excel_display(filename='encryption_keys.xlsx'):
    """
    Fix Excel display issues:
    - Auto-adjust column widths
    - Enable text wrapping for long content
    - Set row heights appropriately
    """
    
    try:
        # Load the workbook
        wb = load_workbook(filename)
        ws = wb.active
        
        print(f"📝 Fixing display for: {filename}")
        
        # Set column widths
        column_widths = {
            'A': 20,  # Username
            'B': 20,  # DES_Key
            'C': 35,  # AES_Key
            'D': 100, # RSA_Private_Key (wide column)
            'E': 60   # Encrypted_Password
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Enable text wrapping for all cells and align top
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='top',
                    horizontal='left'
                )
        
        # Make header row bold and centered
        for cell in ws[1]:
            cell.alignment = Alignment(
                wrap_text=False,
                vertical='center',
                horizontal='center'
            )
            from openpyxl.styles import Font
            cell.font = Font(bold=True)
        
        # Auto-adjust row heights (set minimum height)
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 200  # Tall rows for RSA keys
        
        # Save the workbook
        wb.save(filename)
        print(f"✅ Excel display fixed!")
        print(f"\nNow open {filename} and you should see:")
        print("  • Column D (RSA_Private_Key) is now 100 characters wide")
        print("  • Text wrapping is enabled")
        print("  • Rows are taller to show wrapped content")
        print("\n💡 Tip: Double-click the cell to see the full key")
        
    except FileNotFoundError:
        print(f"❌ File not found: {filename}")
        print("Register a user first to create the file.")
    except Exception as e:
        print(f"❌ Error fixing Excel: {e}")

def create_formatted_keys_file():
    """Create a new Excel file with better formatting from scratch"""
    
    try:
        # Read existing data
        df = pd.read_excel('encryption_keys.xlsx')
        
        # Create new file with formatting
        output_file = 'encryption_keys_formatted.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Keys')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Keys']
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 35
            worksheet.column_dimensions['D'].width = 100
            worksheet.column_dimensions['E'].width = 60
            
            # Apply formatting to all cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                worksheet.row_dimensions[cell.row].height = 200
        
        print(f"✅ Created formatted file: {output_file}")
        print("This file has better column widths and text wrapping!")
        
    except Exception as e:
        print(f"❌ Error: {e}")

def main():
    print("\n" + "="*80)
    print("EXCEL DISPLAY FIXER")
    print("="*80)
    print("\nOptions:")
    print("1. Fix existing encryption_keys.xlsx")
    print("2. Create new formatted copy (encryption_keys_formatted.xlsx)")
    print("3. Fix both files")
    
    choice = input("\nEnter choice (1-3): ").strip()
    
    if choice == '1':
        fix_excel_display('encryption_keys.xlsx')
    elif choice == '2':
        create_formatted_keys_file()
    elif choice == '3':
        fix_excel_display('encryption_keys.xlsx')
        create_formatted_keys_file()
    else:
        print("Invalid choice")

if __name__ == "__main__":
    main()
